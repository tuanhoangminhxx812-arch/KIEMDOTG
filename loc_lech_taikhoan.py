import sys
import os
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import re
import unicodedata


def normalize_header(name):
    if not name:
        return ""
    name = str(name).strip().lower()
    name = name.replace('đ', 'd')
    name = unicodedata.normalize('NFKD', name).encode('ascii', 'ignore').decode('utf-8')
    name = re.sub(r'[^a-z0-9]', '', name)
    return name

def detect_sheet_and_headers(wb):
    header_patterns = {
        'account': ['taikhoan', 'tk'],
        'creator': ['nguoitao', 'nguoilap'],
        'debit': ['noquydoi', 'noquyquy', 'no'],
        'credit': ['coquydoi', 'coquyquy', 'co'],
        'updater': ['nguoicapnhat', 'nguoisua'],
        'trans_num': ['sogiaodich', 'sochungtu', 'soctu', 'sohieuchungtu', 'sohieu', 'sohieugd'],
        'desc': ['noidung', 'diengiai', 'noidunggiaodich', 'noidunggd'],
        'date': ['ngaygiaodich', 'ngaygd', 'ngay', 'ngaychungtu']
    }
    
    for sheet_name in wb.sheetnames:
        if sheet_name in ['DoiChieu_TKTG', 'ChiTiet_GiaoDich_Lech', 'Data_Unpaired']:
            continue
        sheet = wb[sheet_name]
        for row_idx, row in enumerate(sheet.iter_rows(max_row=20, values_only=True), start=1):
            row_normalized = [normalize_header(cell) for cell in row if cell is not None]
            if not row_normalized:
                continue
            has_acc = any(pat in row_normalized for pat in header_patterns['account'])
            if not has_acc:
                continue
                
            detected_indices = {}
            for col_key, patterns in header_patterns.items():
                matched_idx = None
                for pat in patterns:
                    for i, val in enumerate(row_normalized):
                        if val == pat:
                            matched_idx = i
                            break
                        if len(pat) > 2 and pat in val:
                            matched_idx = i
                            break
                    if matched_idx is not None:
                        break
                detected_indices[col_key] = matched_idx
            
            critical_keys = ['account', 'debit', 'credit', 'desc', 'creator']
            if all(detected_indices.get(k) is not None for k in critical_keys):
                # Kiểm tra cột Tài khoản xem có dữ liệu hay không (tránh sheet rỗng hoặc sheet tóm tắt phụ)
                acc_col_idx = detected_indices['account']
                has_values = False
                for r_row in sheet.iter_rows(min_row=row_idx + 1, max_row=row_idx + 100, max_col=acc_col_idx + 1, values_only=True):
                    if len(r_row) > acc_col_idx and r_row[acc_col_idx] is not None and str(r_row[acc_col_idx]).strip() != "":
                        has_values = True
                        break
                if has_values:
                    return sheet_name, row_idx, detected_indices
                
    return None, None, None

def normalize_text(text):
    if not text:
        return ""
    text = text.lower()
    text = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('utf-8')
    text = re.sub(r'[_–\-\/,:;()\[\]_]', ' ', text)
    return " ".join(text.split())

def extract_keywords(text):
    text = normalize_text(text)
    stop_words = {
        'phai', 'tra', 'cty', 'ct', 'theo', 'hoa', 'don', 'ngay', 'thuoc', 
        'mua', 'sam', 'don', 'vi', 'thang', 'cp', 'cho', 'so', 'nho', 'le',
        'chua', 'doi', 'ung', 'co', 'no', 'quy', 'doi', 'scl', 'sxkd'
    }
    words = text.split()
    keywords = [w for w in words if w not in stop_words and len(w) > 1]
    return set(keywords)

def get_similarity(desc1, desc2):
    k1 = extract_keywords(desc1)
    k2 = extract_keywords(desc2)
    if not k1 or not k2:
        return 0.0
    intersection = k1.intersection(k2)
    return len(intersection) / min(len(k1), len(k2))

def are_similar(desc1, desc2, trans1, trans2, threshold=0.35):
    d1_norm = normalize_text(desc1)
    d2_norm = normalize_text(desc2)
    
    # Nếu cả hai diễn giải đều trống/None, coi như tương đồng (đối ứng động ERP)
    if not d1_norm and not d2_norm:
        return True
        
    t1_norm = normalize_text(str(trans1 or ''))
    t2_norm = normalize_text(str(trans2 or ''))
    
    t1_clean = re.sub(r'[^a-zA-Z0-9]', '', t1_norm).lower()
    t2_clean = re.sub(r'[^a-zA-Z0-9]', '', t2_norm).lower()
    
    d1_clean = re.sub(r'[^a-zA-Z0-9]', '', d1_norm).lower()
    d2_clean = re.sub(r'[^a-zA-Z0-9]', '', d2_norm).lower()
    
    # 1. Khớp số giao dịch/số chứng từ trực tiếp hoặc nằm trong nội dung
    if len(t1_clean) > 5 and (t1_clean in d2_clean or t1_clean in d1_clean):
        return True
    if len(t2_clean) > 5 and (t2_clean in d1_clean or t2_clean in d2_clean):
        return True
        
    # 2. Khớp các số từ 3 chữ số trở lên (như số hóa đơn, số hợp đồng), loại bỏ các năm để tránh trùng lặp giả
    nums1 = set(re.findall(r'\d{3,}', d1_norm + ' ' + t1_norm))
    nums2 = set(re.findall(r'\d{3,}', d2_norm + ' ' + t2_norm))
    common_years = {'2024', '2025', '2026', '2027'}
    nums1 = nums1 - common_years
    nums2 = nums2 - common_years
    if nums1.intersection(nums2):
        return True
        
    # 3. Hợp lý hóa so khớp từ khóa đặc thù chuyên ngành
    words1 = set(d1_norm.split())
    words2 = set(d2_norm.split())
    
    # Nhóm Bảo hiểm (BHXH, BHTN, BHYT, KPCD)
    insurance_keywords = {'bhxh', 'bhtn', 'bhyt', 'kpcd', 'bao', 'hiem'}
    if (insurance_keywords.intersection(words1) and insurance_keywords.intersection(words2)):
        return True
        
    # Nhóm Xăng dầu / Nhiên liệu
    fuel_keywords = {'xang', 'dau', 'nhien', 'lieu'}
    if (fuel_keywords.intersection(words1) and fuel_keywords.intersection(words2)):
        return True
        
    # Nhóm Điện
    elec_keywords = {'dien'}
    if (elec_keywords.intersection(words1) and elec_keywords.intersection(words2)):
        return True
        
    # Nhóm Nước
    water_keywords = {'nuoc'}
    if (water_keywords.intersection(words1) and water_keywords.intersection(words2)):
        return True
        
    # Nhóm Viễn thông / Cước / Internet / Xe công nghệ (Grab, Xanh SM)
    telecom_keywords = {'internet', 'cuoc', 'dt', 'dienthoai', 'sm', 'xanh'}
    if (telecom_keywords.intersection(words1) and telecom_keywords.intersection(words2)):
        return True
        
    # Nhóm Kiểm toán / Báo cáo quyết toán
    audit_keywords = {'kiem', 'toan', 'aasc', 'bcktkt'}
    if (audit_keywords.intersection(words1) and audit_keywords.intersection(words2)):
        return True
        
    # Nhóm Điều chuyển / Điều động vật tư nội bộ giữa các đơn vị
    # ("Xuất điều động cho Công ty ĐL..." <-> "Xuất điều chuyển vật tư...")
    dieu_keywords = {'dieu', 'xuat'}
    if (dieu_keywords.issubset(words1) and dieu_keywords.issubset(words2)):
        # Kiểm tra có ít nhất 1 mã đơn vị chung (PCSG, PCTT, PCBC, PCHM, PCVT, DVDL, LDDP, PCGD...)
        unit_codes = {'pcsg', 'pctt', 'pcbc', 'pchm', 'pcvt', 'dvdl', 'ldpp', 'pcgd', 'pccc', 'pccl', 'dvdt'}
        units1 = unit_codes.intersection(words1)
        units2 = unit_codes.intersection(words2)
        if units1 and units2 and units1.intersection(units2):
            return True
        
    # Nhóm Thu tiền bảo lãnh (BL) - hợp đồng ngắn hạn thu hộ
    # Một bên hạch toán nợ PE/AP/GK..., bên kia có Y...
    bl_keywords = {'thu', 'tien', 'bl'}
    if (bl_keywords.issubset(words1) and bl_keywords.issubset(words2)):
        # Cả 2 đều là thu tiền bảo lãnh → kiểm tra số PE/GK trùng nhau
        pe_nums1 = set(re.findall(r'pe\d{8,}|gk\d{5,}', d1_norm.lower()))
        pe_nums2 = set(re.findall(r'pe\d{8,}|gk\d{5,}', d2_norm.lower()))
        if pe_nums1 and pe_nums2 and pe_nums1.intersection(pe_nums2):
            return True
        
    # 4. Tính toán Jaccard similarity trên tập từ khóa
    sim = get_similarity(desc1, desc2)
    return sim >= threshold

def find_subset_sum(numbers, target, max_size=6, tolerance=0.5):
    """
    Tìm một tập hợp con các số có tổng bằng target (trong khoảng sai số tolerance).
    numbers là danh sách các tuple: (idx, value)
    """
    sorted_numbers = sorted(numbers, key=lambda x: x[1], reverse=True)
    n = len(sorted_numbers)
    
    def backtrack(start, current_subset, current_sum):
        if abs(current_sum - target) <= tolerance:
            return current_subset
        if len(current_subset) >= max_size or current_sum > target + tolerance:
            return None
        for i in range(start, n):
            idx, val = sorted_numbers[i]
            res = backtrack(i + 1, current_subset + [idx], current_sum + val)
            if res is not None:
                return res
        return None
        
    return backtrack(0, [], 0.0)


def main(input_path=None):
    sys.stdout.reconfigure(encoding='utf-8')
    
    if input_path is not None:
        file_path = os.path.abspath(input_path)
        base_dir = os.path.dirname(file_path)
        base_name = os.path.basename(file_path)
        name_without_ext, ext = os.path.splitext(base_name)
        backup_path = os.path.join(base_dir, f"{name_without_ext}_backup{ext}")
    else:
        file_path = r'c:\Users\tuan2hm\Downloads\Linh Tinh\GL 0903 T4.2026_lech_08052026.xlsx'
        backup_path = r'c:\Users\tuan2hm\Downloads\Linh Tinh\GL 0903 T4.2026_lech_08052026_backup.xlsx'
    
    # 1. Quản lý File Backup để tránh mất cache công thức của Excel gốc
    if not os.path.exists(backup_path):
        if not os.path.exists(file_path):
            print(f"Không tìm thấy file Excel tại: {file_path}")
            return
        print(f"Đang tạo bản sao lưu dữ liệu gốc lần đầu tiên tại: {backup_path}...")
        shutil.copy2(file_path, backup_path)
    else:
        print("Đã phát hiện file backup dữ liệu gốc. Tiến hành đọc dữ liệu từ file backup...")

    # 2. Đọc dữ liệu từ file backup
    print("Đang đọc dữ liệu gốc...")
    wb_read = openpyxl.load_workbook(backup_path, data_only=True)
    
    source_sheet_name, header_row, indices = detect_sheet_and_headers(wb_read)
    if not source_sheet_name:
        print("Không tìm thấy sheet dữ liệu gốc hợp lệ hoặc lỗi cấu trúc cột trong file Excel.")
        return
        
    sheet = wb_read[source_sheet_name]
    
    acc_idx = indices['account']
    creator_idx = indices['creator']
    deb_idx = indices['debit']
    cred_idx = indices['credit']
    upd_idx = indices['updater']
    trans_idx = indices['trans_num']
    desc_idx = indices['desc']
    date_idx = indices['date']
    
    # Calculate column letters dynamically for the SUMIF formulas in DoiChieu_TKTG sheet
    acc_col = get_column_letter(acc_idx + 1)
    deb_col = get_column_letter(deb_idx + 1)
    cred_col = get_column_letter(cred_idx + 1)
        
    target_prefixes = ('33191', '33192', '33193', '33194', '33196', '33198', '2419', '1510', '1511', '1519')
    intermediate_accounts = [
        '33191000000',
        '33192000000',
        '33193000000',
        '33194000000',
        '33195000000',
        '33196100000',
        '33198000000',
        '24190000000',
        '15110000000',
        '15190000000'
    ]
    
    print("Đang phân tích dữ liệu...")
    
    # Gom dữ liệu theo Tài khoản và Người tạo
    results = {}
    all_creators = set()
    
    for row_num, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not row[acc_idx]:
            continue
        acc = str(row[acc_idx]).strip()
        prefix = next((p for p in target_prefixes if acc.startswith(p)), None)
        if not prefix:
            continue
            
        deb_val = row[deb_idx] or 0
        cred_val = row[cred_idx] or 0
        creator = row[creator_idx] or 'UNKNOWN'
        updater = row[upd_idx] or 'UNKNOWN'
        all_creators.add(creator)
        
        # Chuyển đổi hạch toán âm (Negative Reversals)
        deb = 0
        cred = 0
        if deb_val > 0:
            deb = deb_val
        elif deb_val < 0:
            cred = -deb_val
            
        if cred_val > 0:
            cred = cred_val
        elif cred_val < 0:
            deb = -cred_val
            
        if deb == 0 and cred == 0:
            continue
            
        if acc not in results:
            results[acc] = {}
        if creator not in results[acc]:
            results[acc][creator] = []
            
        results[acc][creator].append({
            'row_num': row_num,
            'deb': deb,
            'cred': cred,
            'creator': creator,
            'updater': updater,
            'trans_num': row[trans_idx],
            'date': row[date_idx],
            'desc': row[desc_idx]
        })
        
    print("Đang chạy thuật toán đối ứng đa cấp (Creator-Level & Transaction-Level Pairing)...")
    unpaired_transactions = {}
    
    for acc in sorted(results.keys()):
        # Tính chênh lệch tổng hợp của tài khoản để xem có lệch không
        all_acc_rows = []
        for c, rows in results[acc].items():
            all_acc_rows.extend(rows)
            
        total_deb = sum(r['deb'] for r in all_acc_rows)
        total_cred = sum(r['cred'] for r in all_acc_rows)
        
        # Nếu tài khoản tự đối ứng khớp hoàn toàn tổng thể -> Bỏ qua không cần phân tích sâu
        if abs(round(total_deb - total_cred, 2)) == 0:
            continue
            
        # 1. PHÂN TÍCH VÀ GHÉP CẶP ĐỐI ỨNG CẤP ĐỘ NGƯỜI TẠO (CREATOR LEVEL DIFFERENCE PAIRING)
        creator_diffs = {}
        for creator, rows in results[acc].items():
            diff = sum(r['deb'] for r in rows) - sum(r['cred'] for r in rows)
            if abs(round(diff, 2)) != 0:
                creator_diffs[creator] = diff
                
        pos_creators = {c: diff for c, diff in creator_diffs.items() if diff > 0.01}
        neg_creators = {c: diff for c, diff in creator_diffs.items() if diff < -0.01}
        
        eliminated_creators = set()
        
        # Bước A: Ghép cặp Creator 1-1 đối xứng chính xác số tiền lệch
        for pc, p_diff in list(pos_creators.items()):
            for nc, n_diff in list(neg_creators.items()):
                if pc not in eliminated_creators and nc not in eliminated_creators:
                    if abs(p_diff + n_diff) < 0.5:
                        eliminated_creators.add(pc)
                        eliminated_creators.add(nc)
                        print(f"  [{acc}] Người tạo {pc} (+{p_diff:,.0f}) và {nc} ({n_diff:,.0f}) tự bù trừ cho nhau.")
                        break
                        
        # Bước B: Ghép cặp Creator 1-nhiều
        for nc, n_diff in list(neg_creators.items()):
            if nc not in eliminated_creators:
                active_pos = [(c, diff) for c, diff in pos_creators.items() if c not in eliminated_creators]
                active_pos_for_subset = [(i, diff) for i, (c, diff) in enumerate(active_pos)]
                subset_indices = find_subset_sum(active_pos_for_subset, -n_diff, max_size=5)
                if subset_indices is not None:
                    eliminated_creators.add(nc)
                    for idx in subset_indices:
                        name, diff = active_pos[idx]
                        eliminated_creators.add(name)
                    print(f"  [{acc}] Nhóm bù trừ 1-nhiều thành công cho {nc}.")
                    
        # Bước C: Ghép cặp Creator nhiều-1
        for pc, p_diff in list(pos_creators.items()):
            if pc not in eliminated_creators:
                active_neg = [(c, -diff) for c, diff in neg_creators.items() if c not in eliminated_creators]
                active_neg_for_subset = [(i, diff) for i, (c, diff) in enumerate(active_neg)]
                subset_indices = find_subset_sum(active_neg_for_subset, p_diff, max_size=5)
                if subset_indices is not None:
                    eliminated_creators.add(pc)
                    for idx in subset_indices:
                        name, diff = active_neg[idx]
                        eliminated_creators.add(name)
                    print(f"  [{acc}] Nhóm bù trừ nhiều-1 thành công cho {pc}.")
                    
        # 2. CHẠY GHÉP CẶP CHI TIẾT CHÉO NGƯỜI TẠO (CROSS-CREATOR SIMILARITY PAIRING)
        # Trước tiên, đánh matched=True cho tất cả dòng của các creator đã được loại trừ ở cấp creator-level
        for c_elim in eliminated_creators:
            for row in results[acc].get(c_elim, []):
                row['matched'] = True

        all_acc_rows = []
        for c, rows in results[acc].items():
            if c not in eliminated_creators:
                all_acc_rows.extend(rows)
                
        debits = [r for r in all_acc_rows if r['deb'] > 0]
        credits = [r for r in all_acc_rows if r['cred'] > 0]
        
        for d in debits:
            d['matched'] = False
        for c in credits:
            c['matched'] = False
            
        # Bước 2A: Ghép cặp chi tiết 1-1 chéo người tạo theo diễn giải tương đồng
        for d in debits:
            if d['matched']:
                continue
            for c in credits:
                if c['matched']:
                    continue
                if abs(d['deb'] - c['cred']) < 0.5:
                    if are_similar(d['desc'], c['desc'], d['trans_num'], c['trans_num']):
                        d['matched'] = True
                        c['matched'] = True
                        break
                        
        # Bước 2B: Ghép cặp chi tiết 1-nhiều chéo người tạo
        for d in debits:
            if d['matched']:
                continue
            similar_credits = [c for c in credits if not c['matched'] and are_similar(d['desc'], c['desc'], d['trans_num'], c['trans_num'])]
            if not similar_credits:
                continue
            target = d['deb']
            
            # --- CẢI TIẾN THUẬT TOÁN CHO TẬP HỢP LỚN ---
            # 1. Kiểm tra nếu tổng tất cả các credit tương đồng khớp với target
            total_similar_sum = sum(c['cred'] for c in similar_credits)
            if abs(total_similar_sum - target) < 0.5:
                d['matched'] = True
                for c in similar_credits:
                    c['matched'] = True
                continue
                
            # 2. Kiểm tra nếu tổng trừ đi 1 phần tử khớp với target
            matched_by_removing_one = False
            for i, c_remove in enumerate(similar_credits):
                if abs(total_similar_sum - c_remove['cred'] - target) < 0.5:
                    d['matched'] = True
                    for j, c in enumerate(similar_credits):
                        if j != i:
                            c['matched'] = True
                    matched_by_removing_one = True
                    break
            if matched_by_removing_one:
                continue
                
            # 3. Kiểm tra nếu tổng trừ đi 2 phần tử khớp với target
            matched_by_removing_two = False
            diff_target = total_similar_sum - target
            if diff_target > 0:
                seen_vals = {}
                for idx, c in enumerate(similar_credits):
                    val = c['cred']
                    needed = diff_target - val
                    found_idx = None
                    for key_val, key_idx in seen_vals.items():
                        if abs(key_val - needed) < 0.5:
                            found_idx = key_idx
                            break
                    if found_idx is not None:
                        d['matched'] = True
                        for j, c_match in enumerate(similar_credits):
                            if j != idx and j != found_idx:
                                c_match['matched'] = True
                        matched_by_removing_two = True
                        break
                    seen_vals[val] = idx
            if matched_by_removing_two:
                continue

            # 4. Fallback về tìm subset sum thông thường (giới hạn size nhỏ)
            credit_numbers = [(i, c['cred']) for i, c in enumerate(similar_credits)]
            subset_indices = find_subset_sum(credit_numbers, target, max_size=6)
            if subset_indices is not None:
                d['matched'] = True
                for idx in subset_indices:
                    similar_credits[idx]['matched'] = True
                    
        # Bước 2C: Ghép cặp chi tiết nhiều-1 chéo người tạo
        for c in credits:
            if c['matched']:
                continue
            similar_debits = [d for d in debits if not d['matched'] and are_similar(d['desc'], c['desc'], d['trans_num'], c['trans_num'])]
            if not similar_debits:
                continue
            target = c['cred']
            
            # --- CẢI TIẾN THUẬT TOÁN CHO TẬP HỢP LỚN ---
            # 1. Kiểm tra nếu tổng tất cả các debit tương đồng khớp với target
            total_similar_sum = sum(d['deb'] for d in similar_debits)
            if abs(total_similar_sum - target) < 0.5:
                c['matched'] = True
                for d in similar_debits:
                    d['matched'] = True
                continue
                
            # 2. Kiểm tra nếu tổng trừ đi 1 phần tử khớp với target
            matched_by_removing_one = False
            for i, d_remove in enumerate(similar_debits):
                if abs(total_similar_sum - d_remove['deb'] - target) < 0.5:
                    c['matched'] = True
                    for j, d in enumerate(similar_debits):
                        if j != i:
                            d['matched'] = True
                    matched_by_removing_one = True
                    break
            if matched_by_removing_one:
                continue
                
            # 3. Kiểm tra nếu tổng trừ đi 2 phần tử khớp với target
            matched_by_removing_two = False
            diff_target = total_similar_sum - target
            if diff_target > 0:
                seen_vals = {}
                for idx, d in enumerate(similar_debits):
                    val = d['deb']
                    needed = diff_target - val
                    found_idx = None
                    for key_val, key_idx in seen_vals.items():
                        if abs(key_val - needed) < 0.5:
                            found_idx = key_idx
                            break
                    if found_idx is not None:
                        c['matched'] = True
                        for j, d_match in enumerate(similar_debits):
                            if j != idx and j != found_idx:
                                d_match['matched'] = True
                        matched_by_removing_two = True
                        break
                    seen_vals[val] = idx
            if matched_by_removing_two:
                continue

            # 4. Fallback về tìm subset sum thông thường
            debit_numbers = [(i, d['deb']) for i, d in enumerate(similar_debits)]
            subset_indices = find_subset_sum(debit_numbers, target, max_size=6)
            if subset_indices is not None:
                c['matched'] = True
                for idx in subset_indices:
                    similar_debits[idx]['matched'] = True
                    
        # Kết thúc vòng lặp tài khoản
        pass
        
    # Thu thập tất cả các giao dịch chưa khớp từ mọi tài khoản
    all_unmatched_debits = []
    all_unmatched_credits = []
    for acc, creator_data in results.items():
        for creator, rows in creator_data.items():
            for r in rows:
                if not r.get('matched', False):
                    r['acc'] = acc  # Lưu tài khoản để đối chiếu chéo
                    if r['deb'] > 0:
                        all_unmatched_debits.append(r)
                    elif r['cred'] > 0:
                        all_unmatched_credits.append(r)

    # Chạy đối chiếu chéo 1-1 khác tài khoản
    print("\nĐang chạy đối chiếu chéo tài khoản (Cross-Account Reconciliation)...")
    for d in all_unmatched_debits:
        if d.get('matched', False):
            continue
        for c in all_unmatched_credits:
            if c.get('matched', False):
                continue
            if d['acc'] == c['acc']:
                continue  # Chỉ ghép chéo khác tài khoản
            if abs(d['deb'] - c['cred']) < 0.5:
                if are_similar(d['desc'], c['desc'], d['trans_num'], c['trans_num']):
                    d['matched'] = True
                    c['matched'] = True
                    print(f"  [Khớp Chéo] TK {d['acc']} (Dòng {d['row_num']}, Nợ {d['deb']:,.0f}) <-> TK {c['acc']} (Dòng {c['row_num']}, Có {c['cred']:,.0f})")
                    break

    # Tạo unpaired_transactions từ các giao dịch thực sự chưa khớp sau đối chiếu chéo
    unpaired_transactions = {}
    for acc in sorted(results.keys()):
        all_acc_rows = []
        for creator, rows in results[acc].items():
            all_acc_rows.extend(rows)
        acc_unpaired = [r for r in all_acc_rows if not r.get('matched', False)]
        if acc_unpaired:
            acc_unpaired.sort(key=lambda x: x['row_num'])
            unpaired_transactions[acc] = acc_unpaired
            
    # 3. Tạo Workbook mới dựa trên bản backup nguyên vẹn để ghi đè báo cáo mới
    wb_write = openpyxl.load_workbook(backup_path)
    
    # Xóa các sheet cũ nếu có
    for sh_name in ['DoiChieu_TKTG', 'ChiTiet_GiaoDich_Lech', 'Data_Unpaired']:
        if sh_name in wb_write.sheetnames:
            del wb_write[sh_name]
            
    # ----------------------------------------------------
    # SHEET PHỤ: Data_Unpaired (Chứa dữ liệu thô của giao dịch bị lệch thực tế)
    # ----------------------------------------------------
    ws_data = wb_write.create_sheet(title='Data_Unpaired')
    ws_data.views.sheetView[0].showGridLines = True
    
    # Tiêu đề bảng
    headers_data = ["Tài khoản", "Người tạo", "Nợ quy đổi", "Có quy đổi", "Số giao dịch", "Ngày giao dịch", "Nội dung", "Dòng gốc (Row)", "Người cập nhật"]
    for col_idx, h_text in enumerate(headers_data, start=1):
        cell = ws_data.cell(row=1, column=col_idx, value=h_text)
        cell.font = Font(name='Segoe UI', size=11, bold=True)
        cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid') # Xanh lá nhạt
        cell.border = Border(bottom=Side(style='medium', color='548235'))
        
    data_row = 2
    for acc, txs in unpaired_transactions.items():
        for tx in txs:
            ws_data.cell(row=data_row, column=1, value=acc)
            ws_data.cell(row=data_row, column=2, value=tx['creator'])
            ws_data.cell(row=data_row, column=3, value=tx['deb'])
            ws_data.cell(row=data_row, column=4, value=tx['cred'])
            ws_data.cell(row=data_row, column=5, value=tx['trans_num'])
            c_date = ws_data.cell(row=data_row, column=6, value=tx['date'])
            if tx['date']:
                c_date.number_format = 'yyyy-mm-dd'
            ws_data.cell(row=data_row, column=7, value=tx['desc'])
            ws_data.cell(row=data_row, column=8, value=tx['row_num'])
            ws_data.cell(row=data_row, column=9, value=tx['updater'])
            data_row += 1
            
    # Định dạng cột của sheet Data_Unpaired
    ws_data.column_dimensions['A'].width = 18
    ws_data.column_dimensions['B'].width = 25
    ws_data.column_dimensions['C'].width = 15
    ws_data.column_dimensions['D'].width = 15
    ws_data.column_dimensions['E'].width = 15
    ws_data.column_dimensions['F'].width = 15
    ws_data.column_dimensions['G'].width = 40
    ws_data.column_dimensions['H'].width = 15
    ws_data.column_dimensions['I'].width = 25
            
    # ----------------------------------------------------
    # SHEET 1: DoiChieu_TKTG (Dashboard Đối chiếu Động)
    # ----------------------------------------------------
    ws_dash = wb_write.create_sheet(title='DoiChieu_TKTG')
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Styles tiêu chuẩn corporate
    font_family = 'Segoe UI'
    title_font = Font(name=font_family, size=16, bold=True, color='1B365D')
    section_font = Font(name=font_family, size=12, bold=True, color='1B365D')
    header_font = Font(name=font_family, size=11, bold=True, color='FFFFFF')
    normal_font = Font(name=font_family, size=11, color='000000')
    bold_font = Font(name=font_family, size=11, bold=True, color='000000')
    italic_font = Font(name=font_family, size=9, italic=True, color='595959')
    
    header_fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid') # Navy Blue
    zebra_fill = PatternFill(start_color='F2F5F8', end_color='F2F5F8', fill_type='solid') # Xám nhạt
    accent_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid') # Xanh nhạt accent
    
    red_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid') # Cảnh báo lệch đỏ
    red_font = Font(name=font_family, size=11, bold=True, color='C00000')
    
    green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid') # Cân đối xanh
    green_font = Font(name=font_family, size=11, bold=True, color='375623')
    
    thin_border_side = Side(border_style='thin', color='BFBFBF')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    double_bottom_border = Border(
        left=thin_border_side, right=thin_border_side,
        top=thin_border_side,
        bottom=Side(border_style='double', color='1F497D')
    )
    
    ws_dash['B2'] = "ĐỐI CHIẾU SỐ LIỆU TÀI KHOẢN TRUNG GIAN - ERP"
    ws_dash['B2'].font = title_font
    ws_dash.row_dimensions[2].height = 25
    ws_dash['B3'] = "Phân tích tự động số lệch chênh lệch Nợ/Có và truy tìm Người tạo phát sinh lệch"
    ws_dash['B3'].font = italic_font
    
    # Khung Dropdown chọn tài khoản tại C5
    ws_dash['B5'] = "Chọn tài khoản đối chiếu:"
    ws_dash['B5'].font = bold_font
    ws_dash['B5'].alignment = Alignment(vertical='center')
    
    ws_dash['C5'] = '33191000000' # Tài khoản mặc định ban đầu
    ws_dash['C5'].font = Font(name=font_family, size=12, bold=True, color='002060')
    ws_dash['C5'].fill = accent_fill
    ws_dash['C5'].alignment = Alignment(horizontal='center', vertical='center')
    ws_dash['C5'].border = Border(
        left=Side(border_style='medium', color='1F497D'),
        right=Side(border_style='medium', color='1F497D'),
        top=Side(border_style='medium', color='1F497D'),
        bottom=Side(border_style='medium', color='1F497D')
    )
    ws_dash.row_dimensions[5].height = 28
    
    dv = DataValidation(type="list", formula1=f'"{",".join(intermediate_accounts)}"', allow_blank=True)
    dv.error ='Vui lòng chọn tài khoản có trong danh sách'
    dv.errorTitle = 'Lựa chọn không hợp lệ'
    dv.prompt = 'Chọn tài khoản trung gian từ danh sách dropdown'
    dv.promptTitle = 'Chọn tài khoản'
    ws_dash.add_data_validation(dv)
    dv.add(ws_dash['C5'])
    
    # ----------------------------------------------------
    # BẢNG A: TỔNG HỢP TRẠNG THÁI CÁC TÀI KHOẢN (BÊN TRÁI - Dữ liệu Gốc thực tế)
    # ----------------------------------------------------
    ws_dash['B7'] = "BẢNG TỔNG HỢP PHÁT SINH LỆCH"
    ws_dash['B7'].font = section_font
    
    headers_left = ["Tài khoản", "Tổng Phát sinh Nợ", "Tổng Phát sinh Có", "Chênh lệch (Nợ-Có)", "Trạng thái"]
    for col_idx, h_text in enumerate(headers_left, start=2):
        cell = ws_dash.cell(row=8, column=col_idx, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws_dash.row_dimensions[8].height = 26
    
    # Đọc tổng phát sinh trực tiếp từ dữ liệu thực tế
    account_totals = {}
    for acc in intermediate_accounts:
        acc_debits = sum(r['deb'] for r in results.get(acc, {}).get('ALL_ROWS_DUMMY', [])) # Tính dựa trên tổng hợp thực tế ban đầu
        # Tính tổng debit và credit thực tế ban đầu
        total_deb_init = 0
        total_cred_init = 0
        for creator, rows in results.get(acc, {}).items():
            total_deb_init += sum(r['deb'] for r in rows)
            total_cred_init += sum(r['cred'] for r in rows)
        account_totals[acc] = {'deb': total_deb_init, 'cred': total_cred_init, 'diff': total_deb_init - total_cred_init}
        
    start_row_left = 9
    for idx, acc in enumerate(intermediate_accounts):
        r = start_row_left + idx
        ws_dash.cell(row=r, column=2, value=acc).font = bold_font
        
        c_deb = ws_dash.cell(row=r, column=3, value=f"=SUMIF('{source_sheet_name}'!{acc_col}:{acc_col}, B{r}, '{source_sheet_name}'!{deb_col}:{deb_col})")
        c_deb.number_format = '#,##0'
        c_deb.font = normal_font
        
        c_cred = ws_dash.cell(row=r, column=4, value=f"=SUMIF('{source_sheet_name}'!{acc_col}:{acc_col}, B{r}, '{source_sheet_name}'!{cred_col}:{cred_col})")
        c_cred.number_format = '#,##0'
        c_cred.font = normal_font
        
        c_diff = ws_dash.cell(row=r, column=5, value=f"=C{r}-D{r}")
        c_diff.number_format = '#,##0'
        c_diff.font = bold_font
        
        c_stat = ws_dash.cell(row=r, column=6, value=f'=IF(ROUND(E{r},2)=0, "Khớp", "Lệch số")')
        c_stat.font = bold_font
        c_stat.alignment = Alignment(horizontal='center')
        
        # Tô màu tĩnh dựa trên giá trị gốc ban đầu
        real_diff = account_totals[acc]['diff']
        if abs(round(real_diff, 2)) != 0:
            c_diff.fill = red_fill
            c_diff.font = red_font
            c_stat.fill = red_fill
            c_stat.font = red_font
        else:
            c_stat.fill = green_fill
            c_stat.font = green_font
            
        for c_idx in range(2, 7):
            ws_dash.cell(row=r, column=c_idx).border = thin_border
        ws_dash.row_dimensions[r].height = 20
        
    # Dòng Tổng cộng Bảng Trái
    tot_row_left = start_row_left + len(intermediate_accounts)
    ws_dash.cell(row=tot_row_left, column=2, value="Tổng cộng").font = bold_font
    
    c_tot_deb = ws_dash.cell(row=tot_row_left, column=3, value=f"=SUM(C{start_row_left}:C{tot_row_left-1})")
    c_tot_deb.number_format = '#,##0'
    c_tot_deb.font = bold_font
    
    c_tot_cred = ws_dash.cell(row=tot_row_left, column=4, value=f"=SUM(D{start_row_left}:D{tot_row_left-1})")
    c_tot_cred.number_format = '#,##0'
    c_tot_cred.font = bold_font
    
    c_tot_diff = ws_dash.cell(row=tot_row_left, column=5, value=f"=C{tot_row_left}-D{tot_row_left}")
    c_tot_diff.number_format = '#,##0'
    c_tot_diff.font = bold_font
    
    c_tot_stat = ws_dash.cell(row=tot_row_left, column=6, value=f'=IF(ROUND(E{tot_row_left},2)=0, "Khớp", "Lệch")')
    c_tot_stat.font = bold_font
    c_tot_stat.alignment = Alignment(horizontal='center')
    
    for c_idx in range(2, 7):
        cell = ws_dash.cell(row=tot_row_left, column=c_idx)
        cell.border = double_bottom_border
        cell.fill = zebra_fill
    ws_dash.row_dimensions[tot_row_left].height = 22
    
    # ----------------------------------------------------
    # BẢNG B: CHI TIẾT CHÊNH LỆCH THEO NGƯỜI TẠO (BÊN PHẢI - DDYNAMIC TỪ SHEET DATA_UNPAIRED)
    # ----------------------------------------------------
    ws_dash['H7'] = "BẢNG TRA CỨU NGƯỜI LÀM LỆCH (ĐỘNG THEO Ô CHỌN C5)"
    ws_dash['H7'].font = section_font
    
    headers_right = ["Người tạo (Creator)", "Tổng Phát sinh Nợ", "Tổng Phát sinh Có", "Chênh lệch (Nợ-Có)", "Trạng thái"]
    for col_idx, h_text in enumerate(headers_right, start=8):
        cell = ws_dash.cell(row=8, column=col_idx, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        
    sorted_creators = sorted(list(all_creators))
    start_row_right = 9
    
    for idx, creator in enumerate(sorted_creators):
        r = start_row_right + idx
        c_name = ws_dash.cell(row=r, column=8, value=creator)
        c_name.font = bold_font
        c_name.border = thin_border
        
        # SUMIFS truy vấn trực tiếp từ bảng dữ liệu lệch thô Data_Unpaired!
        c_deb = ws_dash.cell(row=r, column=9, value=f"=SUMIFS(Data_Unpaired!C:C, Data_Unpaired!A:A, $C$5, Data_Unpaired!B:B, H{r})")
        c_deb.number_format = '#,##0'
        c_deb.font = normal_font
        c_deb.border = thin_border
        
        c_cred = ws_dash.cell(row=r, column=10, value=f"=SUMIFS(Data_Unpaired!D:D, Data_Unpaired!A:A, $C$5, Data_Unpaired!B:B, H{r})")
        c_cred.number_format = '#,##0'
        c_cred.font = normal_font
        c_cred.border = thin_border
        
        c_diff = ws_dash.cell(row=r, column=11, value=f"=I{r}-J{r}")
        c_diff.number_format = '#,##0'
        c_diff.font = bold_font
        c_diff.border = thin_border
        
        c_stat = ws_dash.cell(row=r, column=12, value=f'=IF(AND(I{r}=0,J{r}=0), "", IF(ROUND(K{r},2)=0, "Khớp", "Lệch số"))')
        c_stat.font = bold_font
        c_stat.alignment = Alignment(horizontal='center')
        c_stat.border = thin_border
        
        ws_dash.row_dimensions[r].height = 20
        
    # Tổng cộng của Bảng phải
    tot_row_right = start_row_right + len(sorted_creators)
    ws_dash.cell(row=tot_row_right, column=8, value="Tổng cộng").font = bold_font
    
    c_tot_deb_r = ws_dash.cell(row=tot_row_right, column=9, value=f"=SUM(I{start_row_right}:I{tot_row_right-1})")
    c_tot_deb_r.number_format = '#,##0'
    c_tot_deb_r.font = bold_font
    
    c_tot_cred_r = ws_dash.cell(row=tot_row_right, column=10, value=f"=SUM(J{start_row_right}:J{tot_row_right-1})")
    c_tot_cred_r.number_format = '#,##0'
    c_tot_cred_r.font = bold_font
    
    c_tot_diff_r = ws_dash.cell(row=tot_row_right, column=11, value=f"=I{tot_row_right}-J{tot_row_right}")
    c_tot_diff_r.number_format = '#,##0'
    c_tot_diff_r.font = bold_font
    
    c_tot_stat_r = ws_dash.cell(row=tot_row_right, column=12, value=f'=IF(ROUND(K{tot_row_right},2)=0, "Khớp", "Lệch")')
    c_tot_stat_r.font = bold_font
    c_tot_stat_r.alignment = Alignment(horizontal='center')
    
    for c_idx in range(8, 13):
        cell = ws_dash.cell(row=tot_row_right, column=c_idx)
        cell.border = double_bottom_border
        cell.fill = zebra_fill
    ws_dash.row_dimensions[tot_row_right].height = 22
    
    # Hướng dẫn sử dụng
    ws_dash['B22'] = "Hướng dẫn sử dụng:"
    ws_dash['B22'].font = Font(name=font_family, size=11, bold=True, color='1B365D')
    ws_dash['B23'] = "1. Click chọn ô C5, bấm vào dấu mũi tên dropdown hiện lên và chọn tài khoản cần kiểm tra."
    ws_dash['B23'].font = italic_font
    ws_dash['B24'] = "2. Bảng đối chiếu người tạo bên phải (Cột H -> L) sẽ tự động nhảy số tiền tương ứng tài khoản chọn."
    ws_dash['B24'].font = italic_font
    ws_dash['B25'] = "3. Xem bảng chi tiết ở sheet 'ChiTiet_GiaoDich_Lech' để tìm chính xác dòng bị lỗi để điều chỉnh."
    ws_dash['B25'].font = italic_font

    # ----------------------------------------------------
    # SHEET 2: ChiTiet_GiaoDich_Lech (Báo cáo Chi tiết Mismatch thực tế)
    # ----------------------------------------------------
    ws_detail = wb_write.create_sheet(title='ChiTiet_GiaoDich_Lech')
    ws_detail.views.sheetView[0].showGridLines = True
    
    ws_detail['B2'] = "CHI TIẾT CÁC GIAO DỊCH CHƯA ĐỐI ỨNG (GÂY LỆCH SỐ THỰC TẾ)"
    ws_detail['B2'].font = title_font
    ws_detail.row_dimensions[2].height = 25
    
    ws_detail['B3'] = "Báo cáo liệt kê các giao dịch đơn lẻ bị lệch ròng, loại bỏ tất cả các giao dịch đã đối ứng tự động"
    ws_detail['B3'].font = italic_font
    
    headers_detail = ["Dòng gốc (Row)", "Số chứng từ", "Ngày giao dịch", "Số tiền Nợ", "Số tiền Có", "Người tạo (Creator)", "Người cập nhật", "Nội dung giao dịch"]
    
    curr_row = 6
    # Chỉ lấy những TK có chênh lệch thực sự (tổng Nợ ≠ tổng Có trong danh sách chưa khớp)
    def acc_net_diff(acc):
        rows = unpaired_transactions.get(acc, [])
        return round(sum(r['deb'] for r in rows) - sum(r['cred'] for r in rows), 2)

    unbalanced_accs_sorted = sorted([
        acc for acc in unpaired_transactions.keys()
        if acc in intermediate_accounts and acc_net_diff(acc) != 0
    ])
    
    if not unbalanced_accs_sorted:
        ws_detail.cell(row=curr_row, column=2, value="Tuyệt vời! Không phát hiện tài khoản trung gian nào bị lệch số thực tế sau khi ghép cặp.").font = bold_font
    else:
        for acc in unbalanced_accs_sorted:
            # Tiêu đề khối tài khoản
            ws_detail.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=9)
            title_cell = ws_detail.cell(row=curr_row, column=2, value=f"TÀI KHOẢN: {acc} - Danh sách giao dịch làm lệch số thực tế")
            title_cell.font = Font(name=font_family, size=12, bold=True, color='FFFFFF')
            title_cell.fill = header_fill
            title_cell.alignment = Alignment(vertical='center', indent=1)
            ws_detail.row_dimensions[curr_row].height = 26
            curr_row += 1
            
            # Tiêu đề bảng con
            ws_detail.row_dimensions[curr_row].height = 22
            for col_idx, h_text in enumerate(headers_detail, start=2):
                cell = ws_detail.cell(row=curr_row, column=col_idx, value=h_text)
                cell.font = Font(name=font_family, size=10, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            curr_row += 1
            
            # Ghi danh sách giao dịch chưa ghép cặp
            acc_start_row = curr_row
            for idx, tx in enumerate(unpaired_transactions[acc]):
                ws_detail.row_dimensions[curr_row].height = 20
                
                c_row = ws_detail.cell(row=curr_row, column=2, value=tx['row_num'])
                c_row.alignment = Alignment(horizontal='center')
                c_row.font = bold_font
                
                c_trans = ws_detail.cell(row=curr_row, column=3, value=tx['trans_num'])
                c_trans.alignment = Alignment(horizontal='center')
                c_trans.font = normal_font
                
                c_date = ws_detail.cell(row=curr_row, column=4, value=tx['date'])
                c_date.alignment = Alignment(horizontal='center')
                c_date.font = normal_font
                if tx['date']:
                    c_date.number_format = 'yyyy-mm-dd'
                
                c_deb = ws_detail.cell(row=curr_row, column=5, value=tx['deb'])
                c_deb.number_format = '#,##0'
                c_deb.font = normal_font
                
                c_cred = ws_detail.cell(row=curr_row, column=6, value=tx['cred'])
                c_cred.number_format = '#,##0'
                c_cred.font = normal_font
                
                c_creator = ws_detail.cell(row=curr_row, column=7, value=tx['creator'])
                c_creator.font = bold_font
                c_creator.alignment = Alignment(horizontal='left')
                
                c_upd = ws_detail.cell(row=curr_row, column=8, value=tx['updater'])
                c_upd.font = normal_font
                c_upd.alignment = Alignment(horizontal='left')
                
                c_desc = ws_detail.cell(row=curr_row, column=9, value=tx['desc'])
                c_desc.font = normal_font
                c_desc.alignment = Alignment(horizontal='left')
                
                # Zebra vằn
                cell_fill = zebra_fill if idx % 2 == 1 else PatternFill(fill_type=None)
                for col_idx in range(2, 10):
                    cell = ws_detail.cell(row=curr_row, column=col_idx)
                    cell.border = thin_border
                    if cell_fill.fill_type:
                        cell.fill = cell_fill
                        
                curr_row += 1
                
            # Dòng Tổng cộng chưa khớp
            ws_detail.row_dimensions[curr_row].height = 22
            ws_detail.cell(row=curr_row, column=2, value="Tổng cộng chưa khớp").font = bold_font
            ws_detail.cell(row=curr_row, column=2).alignment = Alignment(horizontal='center')
            
            c_tot_d = ws_detail.cell(row=curr_row, column=5, value=f"=SUM(E{acc_start_row}:E{curr_row-1})")
            c_tot_d.number_format = '#,##0'
            c_tot_d.font = bold_font
            
            c_tot_c = ws_detail.cell(row=curr_row, column=6, value=f"=SUM(F{acc_start_row}:F{curr_row-1})")
            c_tot_c.number_format = '#,##0'
            c_tot_c.font = bold_font
            
            ws_detail.cell(row=curr_row, column=7, value="Chênh lệch (Nợ-Có):").font = bold_font
            ws_detail.cell(row=curr_row, column=7).alignment = Alignment(horizontal='right')
            
            c_tot_diff = ws_detail.cell(row=curr_row, column=8, value=f"=E{curr_row}-F{curr_row}")
            c_tot_diff.number_format = '#,##0'
            c_tot_diff.font = red_font
            c_tot_diff.fill = red_fill
            
            for col_idx in range(2, 10):
                cell = ws_detail.cell(row=curr_row, column=col_idx)
                cell.border = double_bottom_border
                cell.fill = zebra_fill
                
            curr_row += 3
            
    # ----------------------------------------------------
    # TỰ ĐỘNG CĂN CHỈNH ĐỘ RỘNG CỘT (AUTO-FIT COLUMNS)
    # ----------------------------------------------------
    for col in ws_dash.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter in ['A', 'G']: 
            ws_dash.column_dimensions[col_letter].width = 3
            continue
        max_len = 0
        for cell in col:
            if cell.row in [2, 3, 22, 23, 24, 25] or (cell.value and str(cell.value).startswith('=')):
                continue
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws_dash.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    ws_dash.column_dimensions['B'].width = 25
    ws_dash.column_dimensions['C'].width = 20
    ws_dash.column_dimensions['D'].width = 18
    ws_dash.column_dimensions['E'].width = 18
    ws_dash.column_dimensions['F'].width = 15
    ws_dash.column_dimensions['H'].width = 25
    ws_dash.column_dimensions['I'].width = 18
    ws_dash.column_dimensions['J'].width = 18
    ws_dash.column_dimensions['K'].width = 18
    ws_dash.column_dimensions['L'].width = 15
    
    for col in ws_detail.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter in ['A']:
            ws_detail.column_dimensions[col_letter].width = 3
            continue
        max_len = 0
        for cell in col:
            if cell.row in [2, 3] or (cell.value and str(cell.value).startswith('=')):
                continue
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws_detail.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    ws_detail.column_dimensions['I'].width = 50
    ws_detail.column_dimensions['B'].width = 15
    ws_detail.column_dimensions['C'].width = 15
    ws_detail.column_dimensions['D'].width = 15
    ws_detail.column_dimensions['E'].width = 18
    ws_detail.column_dimensions['F'].width = 18
    ws_detail.column_dimensions['G'].width = 25
    ws_detail.column_dimensions['H'].width = 25

    # 4. Ghi file Excel với cơ chế fallback phòng khi bị khóa file
    saved_successfully = False
    saved_path = None
    
    base_dir = os.path.dirname(file_path)
    base_name = os.path.basename(file_path)
    name_without_ext, ext = os.path.splitext(base_name)
    
    paths_to_try = [
        file_path,
        os.path.join(base_dir, f"{name_without_ext}_DoiChieu{ext}")
    ]
    
    # Thêm các tên file phụ dự phòng
    for i in range(1, 10):
        paths_to_try.append(os.path.join(base_dir, f"{name_without_ext}_DoiChieu_{i}{ext}"))
        
    for p in paths_to_try:
        try:
            print(f"Đang lưu báo cáo đối chiếu vào: {p}...")
            wb_write.save(p)
            print(f"==> ĐÃ LƯU THÀNH CÔNG TẠI: {p}")
            saved_successfully = True
            saved_path = p
            break
        except PermissionError:
            print(f"[CẢNH BÁO] File đang bị khóa hoặc không có quyền ghi: {p}")
            
    if not saved_successfully:
        print("\n[CỰC KỲ NGUY CẤP] Tất cả các file dự phòng đều bị khóa! Anh vui lòng đóng các file Excel đang mở và chạy lại nhé.")
        
    print("\n=== HOÀN THÀNH XỬ LÝ ===")
    return saved_path

if __name__ == '__main__':
    main()
