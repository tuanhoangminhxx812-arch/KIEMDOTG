import os
import io
import tempfile
import base64
import openpyxl
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
import loc_lech_taikhoan
import shutil

# Cài đặt trang Streamlit
st.set_page_config(
    page_title="Hệ Thống Kiểm Dò Tài Khoản Trung Gian",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Custom CSS để ẩn hoàn toàn các thành phần Streamlit mặc định và hiển thị iframe rộng hết màn hình
st.markdown("""
<style>
    /* Ẩn Header mặc định của Streamlit */
    [data-testid="stHeader"] {
        display: none !important;
        height: 0 !important;
    }
    
    /* Ẩn Footer mặc định */
    footer {
        visibility: hidden !important;
        display: none !important;
        height: 0 !important;
    }
    
    /* Thiết lập lại padding của khối chứa */
    .block-container {
        padding: 0 !important;
        margin: 0 !important;
        max-width: 100% !important;
        height: 100vh !important;
        overflow: hidden !important;
    }
    
    [data-testid="stAppViewContainer"] {
        padding: 0 !important;
    }
    
    [data-testid="stAppViewBlockContainer"] {
        padding: 0 !important;
    }
    
    /* Cấu hình iframe chiếm trọn vẹn màn hình */
    iframe {
        width: 100% !important;
        height: 100vh !important;
        border: none !important;
        display: block !important;
        margin: 0 !important;
        padding: 0 !important;
    }
</style>
""", unsafe_allow_html=True)

# Khai báo Custom Component từ thư mục custom_ui
parent_dir = os.path.dirname(os.path.abspath(__file__))
build_dir = os.path.join(parent_dir, "custom_ui")

# Đối phó với lỗi đường dẫn tiếng Việt trên Windows bằng cách copy thư mục custom_ui vào thư mục tạm thời có tên ASCII sạch
has_non_ascii = any(ord(char) > 127 for char in build_dir) or " " in build_dir
if has_non_ascii:
    temp_custom_ui_dir = os.path.join(tempfile.gettempdir(), "kiemdo_custom_ui")
    try:
        if os.path.exists(temp_custom_ui_dir):
            shutil.rmtree(temp_custom_ui_dir)
        shutil.copytree(build_dir, temp_custom_ui_dir)
        build_dir = temp_custom_ui_dir
    except Exception as e:
        st.warning(f"Không thể copy thư mục custom_ui sang thư mục tạm: {str(e)}")

my_ui = components.declare_component("my_ui", path=build_dir)

# --- HÀM TÍNH TOÁN DỮ LIỆU ĐỒNG BỘ 100% VỚI EXCEL ---
def load_and_calculate_data(file_path, original_filepath):
    # Tìm backup path để đọc dữ liệu gốc có cache công thức
    backup_path = None
    if original_filepath:
        orig_dir = os.path.dirname(original_filepath)
        orig_name = os.path.basename(original_filepath)
        orig_without_ext, orig_ext = os.path.splitext(orig_name)
        candidate = os.path.join(orig_dir, f"{orig_without_ext}_backup{orig_ext}")
        if os.path.exists(candidate):
            backup_path = candidate
            
    source_path = backup_path if backup_path else file_path
    
    wb_source = openpyxl.load_workbook(source_path, data_only=True)
    source_sheet_name, header_row, indices = loc_lech_taikhoan.detect_sheet_and_headers(wb_source)
    if not source_sheet_name:
        wb_source.close()
        raise ValueError("Không tìm thấy sheet dữ liệu gốc hợp lệ hoặc lỗi cấu trúc cột trong file Excel.")
        
    sheet = wb_source[source_sheet_name]
    
    acc_idx = indices['account']
    deb_idx = indices['debit']
    cred_idx = indices['credit']
    creator_idx = indices['creator']
        
    intermediate_accounts = [
        '11310000000', '14190000000', '15110000000', '15190000000', '24190000000',
        '33191000000', '33192000000', '33193000000', '33194000000', '33195000000',
        '33196100000', '33198000000'
    ]
    
    # Khởi tạo Bảng A
    bang_a_data = {acc: {"account": acc, "debit": 0, "credit": 0, "diff": 0, "status": "Khớp"} for acc in intermediate_accounts}
    all_creators = set()
    
    # Duyệt qua các dòng dữ liệu gốc
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not row[acc_idx]:
            continue
        acc = str(row[acc_idx]).strip()
        
        # Thu thập Người tạo (Creator) xuất hiện trong hệ thống
        if len(row) > creator_idx and row[creator_idx] is not None:
            all_creators.add(str(row[creator_idx]).strip())
            
        if acc in bang_a_data:
            deb_val = row[deb_idx] or 0
            cred_val = row[cred_idx] or 0
            
            bang_a_data[acc]["debit"] += deb_val
            bang_a_data[acc]["credit"] += cred_val

    wb_source.close()

    # Tính toán chênh lệch và trạng thái Bảng A
    for acc in intermediate_accounts:
        item = bang_a_data[acc]
        item["diff"] = item["debit"] - item["credit"]
        item["status"] = "Lệch số" if abs(round(item["diff"], 2)) != 0 else "Khớp"

    # 2. Đọc sheet Data_Unpaired lấy giao dịch chưa khớp
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if 'Data_Unpaired' not in wb.sheetnames:
        wb.close()
        raise ValueError("Ứng dụng chưa được Phân Tích Đối Chiếu. Vui lòng bấm nút Phân Tích.")
        
    ws_data = wb['Data_Unpaired']
    unpaired_transactions = []
    
    for row in ws_data.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
            
        date_val = row[5]
        if date_val and hasattr(date_val, 'strftime'):
            date_str = date_val.strftime('%Y-%m-%d')
        else:
            date_str = str(date_val or '')
            
        if len(row) > 1 and row[1]:
            all_creators.add(str(row[1]).strip())
            
        unpaired_transactions.append({
            "account": str(row[0]).strip(),
            "creator": str(row[1] or 'UNKNOWN').strip(),
            "deb": float(row[2] or 0),
            "cred": float(row[3] or 0),
            "trans_num": str(row[4] or ''),
            "date": date_str,
            "desc": str(row[6] or ''),
            "row_num": int(row[7]),
            "updater": str(row[8] or '') if len(row) > 8 else ''
        })
        
    wb.close()
        
    # Nhóm theo Tài khoản và Người tạo cho Bảng B
    bang_b_data = {}
    for acc in intermediate_accounts:
        bang_b_data[acc] = {}
        for creator in all_creators:
            bang_b_data[acc][creator] = {"creator": creator, "debit": 0, "credit": 0, "diff": 0, "status": ""}
        
    for tx in unpaired_transactions:
        acc = tx["account"]
        creator = tx["creator"]
        if acc not in bang_b_data:
            bang_b_data[acc] = {}
        if creator not in bang_b_data[acc]:
            bang_b_data[acc][creator] = {"creator": creator, "debit": 0, "credit": 0, "diff": 0, "status": ""}
            
        bang_b_data[acc][creator]["debit"] += tx["deb"]
        bang_b_data[acc][creator]["credit"] += tx["cred"]

    # Tính chênh lệch và trạng thái Bảng B
    for acc in intermediate_accounts:
        for creator, item in bang_b_data[acc].items():
            item["diff"] = item["debit"] - item["credit"]
            if abs(round(item["diff"], 2)) != 0:
                item["status"] = "Lệch số"
            elif item["debit"] > 0 or item["credit"] > 0:
                item["status"] = "Khớp"
            else:
                item["status"] = ""
            
    # Chuyển đổi Bảng B sang dạng danh sách sắp xếp theo tên người tạo
    bang_b_list = {
        acc: sorted(list(creators.values()), key=lambda x: x["creator"]) 
        for acc, creators in bang_b_data.items()
    }
    
    # Phân nhóm các giao dịch chi tiết
    chi_tiet_data = {}
    for acc in intermediate_accounts:
        chi_tiet_data[acc] = []
        
    for tx in unpaired_transactions:
        acc = tx["account"]
        if acc in chi_tiet_data:
            chi_tiet_data[acc].append(tx)
            
    for acc in intermediate_accounts:
        chi_tiet_data[acc].sort(key=lambda x: x["row_num"])
        
    # Tính toán KPI
    total_unbalanced_accs = sum(1 for acc in intermediate_accounts if bang_a_data[acc]["status"] == "Lệch số")
    total_unpaired_debit = sum(tx["deb"] for tx in unpaired_transactions)
    total_unpaired_credit = sum(tx["cred"] for tx in unpaired_transactions)
    total_discrepancy = total_unpaired_debit - total_unpaired_credit
    
    kpi_cards = {
        "total_unbalanced_accs": total_unbalanced_accs,
        "total_unpaired_debit": total_unpaired_debit,
        "total_unpaired_credit": total_unpaired_credit,
        "total_discrepancy": total_discrepancy
    }
    
    return {
        "kpi_cards": kpi_cards,
        "doi_chieu_bang_a": list(bang_a_data.values()),
        "doi_chieu_bang_b": bang_b_list,
        "chi_tiet_giao_dich": chi_tiet_data
    }

# --- HÀM TẠO EXCEL BÁO CÁO LỆCH CHI TIẾT IN-MEMORY ---
def generate_excel_in_memory(account, creator, analyzed_filepath):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(analyzed_filepath, data_only=True)
    ws_data = wb['Data_Unpaired']
    rows = []
    
    for row in ws_data.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        acc = str(row[0]).strip()
        cre = str(row[1] or 'UNKNOWN').strip()
        
        if acc == account:
            if not creator or cre == creator:
                rows.append({
                    "row_num": int(row[7]),
                    "trans_num": str(row[4] or ''),
                    "date": str(row[5] or ''),
                    "deb": float(row[2] or 0),
                    "cred": float(row[3] or 0),
                    "creator": cre,
                    "updater": str(row[8] or '') if len(row) > 8 else '',
                    "desc": str(row[6] or '')
                })
    wb.close()
    
    wb_export = openpyxl.Workbook()
    ws = wb_export.active
    ws.title = "ChiTiet_Lech"
    ws.views.sheetView[0].showGridLines = True
    
    font_family = 'Segoe UI'
    title_font = Font(name=font_family, size=14, bold=True, color='1B365D')
    sub_font = Font(name=font_family, size=11, italic=True, color='595959')
    header_font = Font(name=font_family, size=11, bold=True, color='FFFFFF')
    bold_font = Font(name=font_family, size=11, bold=True, color='000000')
    normal_font = Font(name=font_family, size=11, color='000000')
    red_font = Font(name=font_family, size=11, bold=True, color='C00000')
    
    header_fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
    zebra_fill = PatternFill(start_color='F2F5F8', end_color='F2F5F8', fill_type='solid')
    red_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    
    thin_border_side = Side(border_style='thin', color='BFBFBF')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    double_bottom_border = Border(
        left=thin_border_side, right=thin_border_side,
        top=thin_border_side,
        bottom=Side(border_style='double', color='1F497D')
    )
    
    ws['A1'] = "BÁO CÁO CHI TIẾT CÁC GIAO DỊCH CHƯA ĐỐI ỨNG GÂY LỆCH"
    ws['A1'].font = title_font
    
    filter_text = f"Tài khoản trung gian: {account}"
    if creator:
        filter_text += f"   |   Người tạo (Creator): {creator}"
    ws['A2'] = filter_text
    ws['A2'].font = Font(name=font_family, size=11, bold=True, color='1B365D')
    
    ws['A3'] = "Trích xuất tự động từ Hệ thống Kiểm dò Tài khoản Trung Gian Điện lực Vũng Tàu"
    ws['A3'].font = sub_font
    
    headers = ["Dòng gốc (Row)", "Số chứng từ", "Ngày giao dịch", "Số tiền Nợ", "Số tiền Có", "Người tạo (Creator)", "Người cập nhật", "Nội dung giao dịch"]
    header_row_num = 5
    ws.row_dimensions[header_row_num].height = 25
    
    for col_idx, h_text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_num, column=col_idx, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
    current_row = 6
    for idx, r_data in enumerate(rows):
        ws.row_dimensions[current_row].height = 20
        
        c_row = ws.cell(row=current_row, column=1, value=r_data['row_num'])
        c_row.alignment = Alignment(horizontal='center')
        c_row.font = bold_font
        c_row.border = thin_border
        
        c_trans = ws.cell(row=current_row, column=2, value=r_data['trans_num'])
        c_trans.alignment = Alignment(horizontal='center')
        c_trans.font = normal_font
        c_trans.border = thin_border
        
        c_date = ws.cell(row=current_row, column=3, value=r_data['date'])
        c_date.alignment = Alignment(horizontal='center')
        c_date.font = normal_font
        c_date.border = thin_border
        
        c_deb = ws.cell(row=current_row, column=4, value=r_data['deb'])
        c_deb.number_format = '#,##0'
        c_deb.font = normal_font
        c_deb.border = thin_border
        
        c_cred = ws.cell(row=current_row, column=5, value=r_data['cred'])
        c_cred.number_format = '#,##0'
        c_cred.font = normal_font
        c_cred.border = thin_border
        
        c_cre = ws.cell(row=current_row, column=6, value=r_data['creator'])
        c_cre.font = bold_font
        c_cre.border = thin_border
        
        c_upd = ws.cell(row=current_row, column=7, value=r_data['updater'])
        c_upd.font = normal_font
        c_upd.border = thin_border
        
        c_desc = ws.cell(row=current_row, column=8, value=r_data['desc'])
        c_desc.font = normal_font
        c_desc.border = thin_border
        
        if idx % 2 == 1:
            for c_idx in range(1, 9):
                ws.cell(row=current_row, column=c_idx).fill = zebra_fill
                
        current_row += 1
        
    if len(rows) > 0:
        ws.row_dimensions[current_row].height = 22
        ws.cell(row=current_row, column=1, value="Tổng cộng chưa khớp").font = bold_font
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=1).border = double_bottom_border
        
        for c_idx in range(2, 4):
            cell = ws.cell(row=current_row, column=c_idx)
            cell.border = double_bottom_border
            cell.fill = zebra_fill
            
        c_tot_deb = ws.cell(row=current_row, column=4, value=f"=SUM(D6:D{current_row-1})")
        c_tot_deb.number_format = '#,##0'
        c_tot_deb.font = bold_font
        c_tot_deb.border = double_bottom_border
        c_tot_deb.fill = zebra_fill
        
        c_tot_cred = ws.cell(row=current_row, column=5, value=f"=SUM(E6:E{current_row-1})")
        c_tot_cred.number_format = '#,##0'
        c_tot_cred.font = bold_font
        c_tot_cred.border = double_bottom_border
        c_tot_cred.fill = zebra_fill
        
        ws.cell(row=current_row, column=6, value="Chênh lệch (Nợ-Có):").font = bold_font
        ws.cell(row=current_row, column=6).alignment = Alignment(horizontal='right')
        ws.cell(row=current_row, column=6).border = double_bottom_border
        ws.cell(row=current_row, column=6).fill = zebra_fill
        
        c_diff = ws.cell(row=current_row, column=7, value=f"=D{current_row}-E{current_row}")
        c_diff.number_format = '#,##0'
        c_diff.font = red_font
        c_diff.fill = red_fill
        c_diff.border = double_bottom_border
        
        ws.cell(row=current_row, column=8).border = double_bottom_border
        ws.cell(row=current_row, column=8).fill = zebra_fill
        
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.row in [1, 2, 3] or (cell.value and str(cell.value).startswith('=')):
                continue
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    ws.column_dimensions['H'].width = 50
    
    # Ghi vào bộ nhớ đệm BytesIO thay vì lưu file vật lý
    output = io.BytesIO()
    wb_export.save(output)
    output.seek(0)
    return output.getvalue()

# --- KHỞI TẠO STATE VÀ XỬ LÝ ĐỐI CHIẾU MÁY THÀNH VIÊN ---
if "ui_state" not in st.session_state:
    st.session_state.ui_state = {
        "uploaded_filename": None,
        "has_analyzed": False,
        "analysis_data": None,
        "download_trigger": None
    }

if "temp_original_filepath" not in st.session_state:
    st.session_state.temp_original_filepath = None
if "temp_backup_filepath" not in st.session_state:
    st.session_state.temp_backup_filepath = None
if "temp_analyzed_filepath" not in st.session_state:
    st.session_state.temp_analyzed_filepath = None

# Render custom component và nhận phản hồi từ iframe
response = my_ui(state=st.session_state.ui_state, key="my_ui_iframe", height=1000)

# Xử lý các sự kiện truyền thông điệp của iframe
if response:
    action = response.get("action")
    
    if action == "upload":
        filename = response.get("filename")
        base64_data = response.get("data")
        
        if filename and base64_data:
            file_bytes = base64.b64decode(base64_data)
            
            # Tạo file tạm thời để chạy đối chiếu chéo
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(file_bytes)
                st.session_state.temp_original_filepath = tmp.name
                
                # Tạo file backup ngay tại thư mục tạm để loc_lech_taikhoan tìm thấy
                base_dir = os.path.dirname(tmp.name)
                base_name = os.path.basename(tmp.name)
                name_without_ext, ext = os.path.splitext(base_name)
                backup_path = os.path.join(base_dir, f"{name_without_ext}_backup{ext}")
                
                with open(backup_path, "wb") as f_backup:
                    f_backup.write(file_bytes)
                
                st.session_state.temp_backup_filepath = backup_path
                
            # Cập nhật state
            st.session_state.ui_state["uploaded_filename"] = filename
            st.session_state.ui_state["has_analyzed"] = False
            st.session_state.ui_state["analysis_data"] = None
            st.session_state.ui_state["download_trigger"] = None
            st.rerun()
            
    elif action == "analyze":
        if st.session_state.temp_original_filepath:
            try:
                # Chạy thuật toán loc_lech_taikhoan
                saved_path = loc_lech_taikhoan.main(st.session_state.temp_original_filepath)
                st.session_state.temp_analyzed_filepath = saved_path
                
                # Tính toán số liệu chênh lệch
                analysis_data = load_and_calculate_data(
                    saved_path,
                    st.session_state.temp_original_filepath
                )
                
                st.session_state.ui_state["has_analyzed"] = True
                st.session_state.ui_state["analysis_data"] = analysis_data
                st.session_state.ui_state["download_trigger"] = None
                st.rerun()
            except Exception as e:
                st.error(f"Lỗi phân tích đối chiếu: {str(e)}")
                
    elif action == "export":
        account = response.get("account")
        creator = response.get("creator")
        
        if account and st.session_state.temp_analyzed_filepath:
            # Tạo báo cáo Excel in-memory
            excel_bytes = generate_excel_in_memory(
                account,
                creator,
                st.session_state.temp_analyzed_filepath
            )
            
            excel_base64 = base64.b64encode(excel_bytes).decode("utf-8")
            
            dl_filename = f"ChiTiet_Lech_{account}"
            if creator:
                dl_filename += f"_{creator}"
            dl_filename += ".xlsx"
            
            # Kích hoạt trigger download phía JS
            st.session_state.ui_state["download_trigger"] = {
                "filename": dl_filename,
                "data": excel_base64
            }
            st.rerun()
            
    elif action == "download_done":
        st.session_state.ui_state["download_trigger"] = None
        st.rerun()
